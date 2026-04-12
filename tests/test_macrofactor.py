"""Tests for MacroFactor xlsx parsing (load_macrofactor_csv in weekly_summary.py)."""

import tempfile
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from scripts.weekly_summary import MACROFACTOR_CORE_COLS, load_macrofactor_csv


def _create_test_xlsx(rows: list[list], sheet_name: str = "Quick Export") -> Path:
    """Create a test xlsx with headers and data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write all 13 core columns as headers
    for i, col in enumerate(MACROFACTOR_CORE_COLS, 1):
        ws.cell(row=1, column=i, value=col)

    # Write data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return Path(tmp.name)


class TestLoadMacrofactorCsv:
    def test_basic_parsing(self):
        xlsx = _create_test_xlsx([
            # Date, Expenditure, Trend Weight, Weight, Calories, Protein, Fat, Carbs,
            # Target Cal, Target Protein, Target Fat, Target Carbs, Steps
            [datetime(2026, 4, 1), 2200, 150.0, 150.5, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
            [datetime(2026, 4, 2), 2300, 149.5, 149.0, 2000, 150, 70, 200, 1900, 145, 65, 190, 10000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        assert len(lines) == 3  # header + 2 data rows
        assert lines[0].strip() == ",".join(MACROFACTOR_CORE_COLS)
        assert "2026-04-01" in lines[1]
        assert "1800" in lines[1]

    def test_date_filtering(self):
        xlsx = _create_test_xlsx([
            [datetime(2026, 3, 28), 2200, 151.0, 151.0, 2100, 130, 70, 210, 1900, 145, 65, 190, 7000],
            [datetime(2026, 4, 1), 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
            [datetime(2026, 4, 8), 2200, 149.0, 149.0, 1900, 145, 62, 185, 1900, 145, 65, 190, 9000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        assert len(lines) == 2  # header + 1 row (only Apr 1 is in range)
        assert "2026-04-01" in lines[1]
        assert "2026-03-28" not in csv_text
        assert "2026-04-08" not in csv_text

    def test_date_as_date_object(self):
        """Dates can be stored as date objects (not datetime) in xlsx."""
        xlsx = _create_test_xlsx([
            [date(2026, 4, 1), 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        assert len(lines) == 2
        assert "2026-04-01" in lines[1]

    def test_none_values_become_empty(self):
        xlsx = _create_test_xlsx([
            [datetime(2026, 4, 1), 2200, 150.0, None, 1800, 140, None, 180, 1900, 145, 65, 190, None],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        # None values should become empty strings in CSV
        parts = lines[1].split(",")
        assert parts[3] == ""  # Weight (lbs) was None
        assert parts[6] == ""  # Fat (g) was None
        assert parts[12] == ""  # Steps was None

    def test_skip_none_date_rows(self):
        xlsx = _create_test_xlsx([
            [datetime(2026, 4, 1), 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
            [None, 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        assert len(lines) == 2  # header + 1 row (None date row skipped)

    def test_empty_window(self):
        xlsx = _create_test_xlsx([
            [datetime(2026, 3, 20), 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        lines = csv_text.strip().split("\n")
        assert len(lines) == 1  # header only

    def test_14_day_window(self):
        """Simulate the real 14-day window (prev_start to this_end)."""
        xlsx = _create_test_xlsx([
            [datetime(2026, 3, 23), 2100, 151.0, 151.0, 1900, 135, 65, 200, 1900, 145, 65, 190, 7500],
            [datetime(2026, 3, 30), 2200, 150.5, 150.5, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
            [datetime(2026, 4, 5), 2300, 149.5, 149.0, 2000, 150, 70, 200, 1900, 145, 65, 190, 9500],
        ])
        # 14-day window: Mar 24 to Apr 6
        csv_text = load_macrofactor_csv(xlsx, date(2026, 3, 24), date(2026, 4, 6))
        lines = csv_text.strip().split("\n")
        # Mar 23 is outside, Mar 30 and Apr 5 are inside
        assert len(lines) == 3  # header + 2 rows
        assert "2026-03-23" not in csv_text
        assert "2026-03-30" in csv_text
        assert "2026-04-05" in csv_text

    def test_all_core_columns_present(self):
        """Verify all 13 core columns are in the output header."""
        xlsx = _create_test_xlsx([
            [datetime(2026, 4, 1), 2200, 150.0, 150.0, 1800, 140, 60, 180, 1900, 145, 65, 190, 8000],
        ])
        csv_text = load_macrofactor_csv(xlsx, date(2026, 4, 1), date(2026, 4, 7))
        header = csv_text.strip().split("\n")[0]
        for col in MACROFACTOR_CORE_COLS:
            assert col in header
